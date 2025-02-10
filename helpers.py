# helpers.py

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

COLORS = {
    # Define the fill color for too low initial concentration
    "initial_fill": PatternFill(start_color="AAAA00", end_color="AAAA00", fill_type="solid"),

    # Define the fill color for too low final concentration
    "final_fill": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),

    # Define the fill color for too low depletion
    "depletion_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),

    # Define the fill color for good difference
    "difference_fill": PatternFill(start_color="3DA564", end_color="3DA564", fill_type="solid"),

    # Define the fill color for good result
    "good_result_fill" : PatternFill(start_color="C7FF87", end_color="C7FF87", fill_type="solid")
}

def detect_target_columns(sheet):
    # Get the maximum number of columns
    max_column = sheet.max_column

    # Initialize a dictionary to store column letters and their headers
    column_headers = {}

    # Iterate over each column in the first row (assumed to be the header row)
    for col in range(1, max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        header = sheet.cell(row=1, column=col).value
        column_headers[header] = column_letter

    target_columns = {
        "Sample Number": column_headers["Sample Number"],
        "Init": column_headers["Init"],
        "Final": column_headers["Final"],
        "Depl": column_headers["Depl"],
        "BOD": column_headers["BOD"],
        "Result": openpyxl.utils.get_column_letter(max_column+1),
        "Average": openpyxl.utils.get_column_letter(max_column+2),
        "Difference": openpyxl.utils.get_column_letter(max_column+3)
    }
    return target_columns


def calculate_average(sample_result):
    average = (sample_result[0] + sample_result[1])/2
    difference = abs(round(((sample_result[0] - sample_result[1])/average * 100), 1))
    return average, difference


def color_row(next_sample, sheet, row):
    if next_sample:
        row_fill = PatternFill(start_color="DD0000", end_color="DD0000", fill_type="none")
    elif  not next_sample:
        row_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = row_fill

def color_bad_final_reading_row(sheet, row, color):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = color

def check_o2_initial(sheet, row):
    pass

# def open_with_libreoffice(file_path):
#         try:
#             # Check the OS and set the command accordingly
#             if os.name == 'nt':  # For Windows
#                 subprocess.run(['soffice', file_path], check=True)
#             else:  # For Unix-like OS including macOS and Linux
#                 subprocess.run(['libreoffice', file_path], check=True)
#             print(f"File {file_path} opened successfully with LibreOffice.")
#         except Exception as e:
#             print(f"Failed to open file {file_path} with LibreOffice. Error: {e}")